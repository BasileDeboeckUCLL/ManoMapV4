# === CLEAN EXPORT.PY (fully rewritten) ===
from typing import Optional, Dict, Any
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import filedialog

# ARGB (A=FF opaque) -> voorkomt 'onzichtbare' fills in Excel
BRIGHT_GREEN = "FF00FF00"  # felgroen
BRIGHT_RED   = "FFFF0000"  # felrood

# --- disabled sections helpers (compat voor sensors.py) ---
_DISABLED_SECTIONS = set()

def add_disabled_sections(items):
    """Voegt 1 of meerdere secties toe aan de 'disabled' set."""
    global _DISABLED_SECTIONS
    if items is None:
        return
    if isinstance(items, (list, tuple, set)):
        _DISABLED_SECTIONS.update(str(x) for x in items)
    else:
        _DISABLED_SECTIONS.add(str(items))

def remove_disabled_sections(items):
    """Verwijdert 1 of meerdere secties uit de 'disabled' set (als ze bestaan)."""
    global _DISABLED_SECTIONS
    if items is None:
        return
    if isinstance(items, (list, tuple, set)):
        for x in items:
            _DISABLED_SECTIONS.discard(str(x))
    else:
        _DISABLED_SECTIONS.discard(str(items))

def reset_disabled_sections():
    """Leegt de 'disabled' set (gebruikt bij reset/back)."""
    global _DISABLED_SECTIONS
    _DISABLED_SECTIONS.clear()

def get_disabled_sections():
    return frozenset(_DISABLED_SECTIONS)
# --- einde helpers ---


# --- kolomhelpers ---
def _get_col(df: pd.DataFrame, candidates):
    """Zoek robuust een kolomnaam in df op basis van kandidaten (case-insensitive)."""
    lower = {c.lower(): c for c in df.columns}
    for name in candidates:
        if name.lower() in lower:
            return lower[name.lower()]
    return None

def _ensure_numeric(s: pd.Series, default=0):
    """Zet naar numeriek; NaN -> default."""
    return pd.to_numeric(s, errors="coerce").fillna(default)


# --- kleurhelpers (fuzzy header match + fallback op HAPC/HARPC + amplitude-only) ---
def _find_col_ix(ws, preferred=None, contains_any=None):
    """
    Zoekt kolomindex (1-based) in header (rij 1).
    - preferred: lijst met exacte namen (case-insensitive)
    - contains_any: lijst met substrings (case-insensitive)
    """
    preferred = preferred or []
    contains_any = contains_any or []
    header_values = [(i+1, (cell.value or "").strip()) for i, cell in enumerate(ws[1])]
    lower_map = {(val.lower()): ix for ix, val in header_values if val}

    # exacte matches eerst
    for name in preferred:
        ix = lower_map.get(name.lower())
        if ix:
            return ix

    # anders substring-matches
    for ix, val in header_values:
        lv = val.lower()
        if any(substr.lower() in lv for substr in contains_any):
            return ix

    return None

def color_statistics_sheet(ws, amp_threshold=100.0,
                           hapc_keys=None, harpc_keys=None,
                           dir_keys=None, amp_keys=None):
    """
    Kleurt rijen in 'Statistics' op basis van:
      1) Direction + amplitude (als beschikbaar)
      2) Fallback A: HARPC/HAPC-counts
      3) Fallback B: alleen amplitude (direction-agnostisch -> groen)
    """
    hapc_keys = (hapc_keys or
                 ["HAPC_count", "hapc_count", "Aantal HAPC", "Aantal HAPCs", "HAPC"])
    harpc_keys = (harpc_keys or
                  ["HARPC_count", "harpc_count", "Aantal HARPC", "Aantal HARPCs", "HARPC"])
    dir_keys = (dir_keys or
                ["direction", "richting", "dir", "pattern_direction", "Direction", "Richting"])
    amp_keys = (amp_keys or
                ["Amplitude_max", "Amplitude_mean", "amplitude", "Amplitude (mmHg)", "mmHg", "Max amplitude"])

    dir_ix   = _find_col_ix(ws, preferred=dir_keys, contains_any=["direction", "richting", "dir"])
    amp_ix   = _find_col_ix(ws, preferred=amp_keys, contains_any=["amplitude", "mmhg"])
    hapc_ix  = _find_col_ix(ws, preferred=hapc_keys, contains_any=["hapc"])
    harpc_ix = _find_col_ix(ws, preferred=harpc_keys, contains_any=["harpc"])

    # debug headers
    try:
        headers_dbg = [(cell.value or "") for cell in ws[1]]
        print("Statistics headers detected:", headers_dbg)
        print("dir_ix:", dir_ix, "amp_ix:", amp_ix, "hapc_ix:", hapc_ix, "harpc_ix:", harpc_ix)
    except Exception:
        pass

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        fill = None

        # 1) Primair: direction + amplitude-drempel
        if dir_ix and amp_ix:
            dval = str(row[dir_ix-1].value).strip().lower()[:1]
            try:
                v = row[amp_ix-1].value
                amp_v = float(v) if v is not None and str(v).strip() != "" else 0.0
            except Exception:
                amp_v = 0.0
            if dval == "a" and amp_v >= float(amp_threshold):
                fill = PatternFill(start_color=BRIGHT_GREEN, end_color=BRIGHT_GREEN, fill_type="solid")
            elif dval == "r" and amp_v >= float(amp_threshold):
                fill = PatternFill(start_color=BRIGHT_RED, end_color=BRIGHT_RED, fill_type="solid")

        # 2) Fallback A: op basis van HARPC/HAPC-counts
        if fill is None and (hapc_ix or harpc_ix):
            try:
                hapc_val  = float(row[hapc_ix-1].value or 0) if hapc_ix else 0.0
                harpc_val = float(row[harpc_ix-1].value or 0) if harpc_ix else 0.0
            except Exception:
                hapc_val = harpc_val = 0.0
            if harpc_val > 0:
                fill = PatternFill(start_color=BRIGHT_RED, end_color=BRIGHT_RED, fill_type="solid")
            elif hapc_val > 0:
                fill = PatternFill(start_color=BRIGHT_GREEN, end_color=BRIGHT_GREEN, fill_type="solid")

        # 3) Fallback B: alleen amplitude (direction onbekend) -> groen
        if fill is None and amp_ix:
            try:
                v = row[amp_ix-1].value
                amp_v = float(v) if v is not None and str(v).strip() != "" else 0.0
            except Exception:
                amp_v = 0.0
            if amp_v >= float(amp_threshold):
                fill = PatternFill(start_color=BRIGHT_GREEN, end_color=BRIGHT_GREEN, fill_type="solid")

        if fill:
            for cell in row:
                cell.fill = fill


def exportToXlsx(df: pd.DataFrame,
                 file_name: str,
                 sliders=None,
                 events=None,
                 settings_sliders=None,
                 first_event_text=None,
                 settings: Optional[Dict[str, Any]] = None):
    """
    Exporteert analyse naar Excel met extra statistiek-tab en kleurcodering.
    Compatibel met bestaande aanroep; ondersteunt optionele `settings` dict.
    """
    # ===== Veiligheidschecks op UI-objecten =====
    if sliders is None or settings_sliders is None:
        print("Export error: sliders/settings_sliders not passed.")
        return

    # Sommige projecten geven widgets door; andere al kale waarden. Afvangen:
    def _get_val(x, default=None):
        try:
            return x.get() if hasattr(x, "get") else x
        except Exception:
            return default

    # ===== Defaults volgens Basile =====
    defaults = {
        "distance_between_sensors": _get_val(settings_sliders[0], 10),  # project gebruikt index 0
        "long_threshold_sensors": 5,
        "hapc_min_sensors": 3,
        "hapc_min_amplitude": 100.0,
    }
    if isinstance(settings, dict):
        defaults.update({k: _get_val(v, defaults.get(k)) for k, v in settings.items()})

    distance_between_sensors = int(round(float(defaults["distance_between_sensors"])) or 10)
    long_thr = int(round(float(defaults["long_threshold_sensors"])) or 5)
    hapc_min_sens = int(round(float(defaults["hapc_min_sensors"])) or 3)
    hapc_min_amp = float(defaults["hapc_min_amplitude"]) or 100.0

    first_event_str = None
    if first_event_text is not None and hasattr(first_event_text, "get"):
        try:
            first_event_str = (first_event_text.get() or "Post-Wake").strip()
        except Exception:
            first_event_str = "Post-Wake"
    if not first_event_str:
        first_event_str = "Post-Wake"

    print("Settings => distance:", distance_between_sensors,
          "long_thr:", long_thr,
          "hapc_min_sens:", hapc_min_sens,
          "hapc_min_amp:", hapc_min_amp,
          "first_event:", first_event_str)

    # ===== Schrijf basis-output =====
    base_name, ext = file_name.rsplit('.', 1)
    new_file_name = f"{base_name}_analysis.xlsx"

    try:
        if df is None or not isinstance(df, pd.DataFrame) or df.empty:
            df = pd.DataFrame({"Info": ["No data provided"], "File": [file_name]})
        df.to_excel(new_file_name, index=False, sheet_name="Sheet1")
    except Exception as e:
        print(f"Error during initial export: {e}")
        return

    # ===== Kolommen bepalen & voorbereiden =====
    print("DEBUG columns:", list(df.columns)[:60])

    direction_col = _get_col(df, ["direction", "richting", "dir", "pattern_direction"])
    longshort_col = _get_col(df, ["longshort", "length_type", "pattern_length", "long_short", "ls", "long/short"])
    region_col    = _get_col(df, ["colonregion", "colon_region", "region", "regio", "colon"])
    event_col     = _get_col(df, ["event", "eventname", "event_name", "event_label"])
    amp_col       = _get_col(df, ["amplitude", "amp", "mmhg"])
    speed_col     = _get_col(df, ["speed", "velocity", "mm/s", "cm/s"])
    senscnt_col   = _get_col(df, ["sensor_count", "sensors", "nsensors", "length_sensors"])

    # Snelheid construeren indien niet aanwezig
    if speed_col is None:
        dist_col = _get_col(df, ["distance_cm", "distance_mm", "distance", "travel_distance"])
        dur_col  = _get_col(df, ["duration_s", "duration", "time_s", "time"])
        if dist_col and dur_col:
            dist_vals = _ensure_numeric(df[dist_col])
            if "cm" in dist_col.lower():
                dist_vals = dist_vals * 10.0  # cm -> mm
            speed_series = dist_vals / _ensure_numeric(df[dur_col], default=np.nan)
            df["__speed_mm_s"] = speed_series.replace([np.inf, -np.inf], np.nan).fillna(0)
            speed_col = "__speed_mm_s"

    # Fallbacks zodat statistiek altijd kan draaien
    needed = {"direction": direction_col, "amplitude": amp_col}
    for k, v in needed.items():
        if v is None:
            print(f"Stats skipped: missing required column for {k}")
            direction_col = direction_col or "direction"
            amp_col = amp_col or "amplitude"
            if direction_col not in df.columns:
                df[direction_col] = "unknown"
            if amp_col not in df.columns:
                df[amp_col] = 0

    # Normaliseren voor grouping
    tmp = df.copy()
    tmp[direction_col] = tmp[direction_col].astype(str).str.lower().str[:1]  # a/r/s
    if longshort_col is None:
        if senscnt_col is None:
            length_col = _get_col(df, ["length", "travel_distance", "distance_cm", "distance_mm"])
            if length_col is not None:
                sens_est = _ensure_numeric(df[length_col]) / max(distance_between_sensors, 1)
                tmp["__longshort"] = np.where(sens_est >= long_thr, "long", "short")
            else:
                tmp["__longshort"] = "short"
        else:
            tmp["__longshort"] = np.where(_ensure_numeric(df[senscnt_col]) >= long_thr, "long", "short")
        longshort_col = "__longshort"
    else:
        tmp[longshort_col] = tmp[longshort_col].astype(str).str.lower().replace({"l": "long", "s": "short"})

    if region_col is None:
        tmp["__region"] = "unknown"
        region_col = "__region"
    if event_col is None:
        tmp["__event"] = first_event_str
        event_col = "__event"
    if speed_col is None:
        tmp["__speed_mm_s"] = 0.0
        speed_col = "__speed_mm_s"

    # Numeriek maken
    tmp[amp_col] = _ensure_numeric(tmp[amp_col])
    tmp[speed_col] = _ensure_numeric(tmp[speed_col])

    # HAPC/HARPC flags (benadering)
    if senscnt_col is None:
        sens_ok = np.where(tmp[longshort_col].astype(str).str.lower().eq("long"), True, False)
    else:
        sens_ok = _ensure_numeric(tmp[senscnt_col]) >= hapc_min_sens
    dir_a = tmp[direction_col].eq("a")
    dir_r = tmp[direction_col].eq("r")
    tmp["__HAPC"]  = (dir_a & sens_ok & (tmp[amp_col] >= hapc_min_amp)).astype(int)
    tmp["__HARPC"] = (dir_r & sens_ok & (tmp[amp_col] >= hapc_min_amp)).astype(int)

    # ===== 1) Statistics & Totals schrijven met pandas =====
    stats = (tmp
             .groupby([direction_col, longshort_col, region_col, event_col], dropna=False)
             .agg(
                 Amplitude_mean=(amp_col, "mean"),
                 Amplitude_max=(amp_col, "max"),
                 Amplitude_min=(amp_col, "min"),
                 Speed_mean=(speed_col, "mean"),
                 Speed_max=(speed_col, "max"),
                 Speed_min=(speed_col, "min"),
                 Count=(amp_col, "size"),
                 HAPC_count=("__HAPC", "sum"),
                 HARPC_count=("__HARPC", "sum"),
             )
             .reset_index()
             .fillna(0)
             )

    totals_dir = (tmp
                  .groupby([direction_col], dropna=False)
                  .agg(
                      Amplitude_mean=(amp_col, "mean"),
                      Amplitude_max=(amp_col, "max"),
                      Amplitude_min=(amp_col, "min"),
                      Speed_mean=(speed_col, "mean"),
                      Speed_max=(speed_col, "max"),
                      Speed_min=(speed_col, "min"),
                      Count=(amp_col, "size"),
                      HAPC_count=("__HAPC", "sum"),
                      HARPC_count=("__HARPC", "sum"),
                  )
                  .reset_index()
                  .fillna(0))

    try:
        with pd.ExcelWriter(new_file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            stats.to_excel(writer, sheet_name="Statistics", index=False)
        with pd.ExcelWriter(new_file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            totals_dir.to_excel(writer, sheet_name="Stats_Totals_by_direction", index=False)
    except Exception as e:
        print(f"Error writing statistics sheets: {e}")
        # return  # desgewenst afbreken

    # ===== 2) Post-processing met openpyxl (headers + kleuren + save) =====
    try:
        wb = load_workbook(new_file_name)

        # Kolomkop 'aantal' opruimen in alle sheets
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for cell in ws[1]:
                if cell.value and isinstance(cell.value, str) and "aantal" in cell.value.lower():
                    cell.value = (cell.value
                                  .replace("aantal ", "")
                                  .replace("Aantal ", "")
                                  .replace("aantal", "")
                                  .strip())

        # Statistics-blad kleuren (robust + fallback)
        if "Statistics" in wb.sheetnames:
            ws_stats = wb["Statistics"]
            color_statistics_sheet(ws_stats, amp_threshold=hapc_min_amp)

        # (optioneel) hoofdblad kleuren alleen als die headers heeft die kloppen
        # try:
        #     ws_main = wb[wb.sheetnames[0]]
        #     color_statistics_sheet(ws_main, amp_threshold=hapc_min_amp)
        # except Exception:
        #     pass

        # Save-as dialoog
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=new_file_name
        )
        if save_path:
            wb.save(save_path)
            print(f"Data successfully exported to {save_path}")
        else:
            wb.save(new_file_name)
            print(f"Data successfully exported to {new_file_name} (default path)")

    except Exception as e:
        print(f"Error exporting data to Excel: {e}")
# === END CLEAN EXPORT.PY ===
