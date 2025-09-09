from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from itertools import chain

# Global constants
EVENT_COLOR = "F0FC5A"
disabled_sections = []
HIGH_AMPLITUDE_MINIMUM_VALUE = 100
HIGH_AMPLITUDE_MINIMUM_PATTERN_LENGTH = 3
LONG_PATTERN_MINIMUM_SENSORS = 5
ALTERNATING_EVENT_COLORS = ["F0FC5A", "FDE9D9"]

def get_pattern_parameters(pattern_params):
    """Extract pattern parameters from GUI inputs with defaults"""
    try:
        long_sensors = int(pattern_params['long_sensors'].get() or 5)
        hapc_sensors = int(pattern_params['hapc_sensors'].get() or 5)
        hapc_consecutive = int(pattern_params['hapc_consecutive'].get() or 3)
        hapc_amplitude = int(pattern_params['hapc_amplitude'].get() or 100)
        
        return {
            'LONG_PATTERN_MINIMUM_SENSORS': long_sensors,
            'HAPC_PATTERN_MINIMUM_SENSORS': hapc_sensors,
            'HIGH_AMPLITUDE_MINIMUM_PATTERN_LENGTH': hapc_consecutive,
            'HIGH_AMPLITUDE_MINIMUM_VALUE': hapc_amplitude
        }
    except (ValueError, TypeError):
        # Return defaults if parsing fails
        return {
            'LONG_PATTERN_MINIMUM_SENSORS': 5,
            'HAPC_PATTERN_MINIMUM_SENSORS': 5,
            'HIGH_AMPLITUDE_MINIMUM_PATTERN_LENGTH': 3,
            'HIGH_AMPLITUDE_MINIMUM_VALUE': 100
        }

def initialize_comprehensive_statistics(event_names):
    """Initialize the comprehensive statistics structure for the new table"""
    regions = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    pattern_types = ["Long s", "Short s", "Long r", "Short r", "Long a", "Short a", "Long ", "Short "]
    region_ranges = ["Ascending - Rectum", "Transverse - Rectum", "Descending - Rectum", "Sigmoid-Rectum"]
    
    comprehensive_stats = {}
    
    for event in event_names:
        comprehensive_stats[event] = {}
        
        for pattern_type in pattern_types:
            comprehensive_stats[event][pattern_type] = {}
            
            for region in regions:
                comprehensive_stats[event][pattern_type][region] = {
                    'count': 0, 'velocities': [], 'amplitudes': []
                }
            
            for region_range in region_ranges:
                comprehensive_stats[event][pattern_type][region_range] = {
                    'count': 0, 'velocities': [], 'amplitudes': []
                }
            
            comprehensive_stats[event][pattern_type]['Total'] = {
                'count': 0, 'velocities': [], 'amplitudes': []
            }
        
        # Special categories
        for special in ['cyclic s', 'cyclic r', 'cyclic a', 'HAPCs', 'HARPCs']:
            comprehensive_stats[event][special] = {'count': 0, 'velocities': [], 'amplitudes': []}
    
    return comprehensive_stats

def classify_pattern_enhanced(row, sliders, distance_between_sensors, params=None):
    """Enhanced pattern classification with new rules and error handling"""
    if params is None:
        params = {
            'LONG_PATTERN_MINIMUM_SENSORS': 5,
            'HAPC_PATTERN_MINIMUM_SENSORS': 5,
            'HIGH_AMPLITUDE_MINIMUM_PATTERN_LENGTH': 3,
            'HIGH_AMPLITUDE_MINIMUM_VALUE': 100
        }

    try:
        length_sensors = 0
        if len(row) > 9 and row[9] and row[9].value is not None:
            try:
                length_sensors = int(float(row[9].value))
            except (ValueError, TypeError):
                length_sensors = 0
        
        direction = ''
        if len(row) > 5 and row[5] and row[5].value is not None:
            direction = str(row[5].value).strip()
            if direction not in ['a', 'r', 's']:
                for char in direction.lower():
                    if char in ['a', 'r', 's']:
                        direction = char
                        break
                else:
                    direction = ''
        
        velocity = 0
        if len(row) > 6 and row[6] and row[6].value is not None:
            try:
                velocity = float(row[6].value)
            except (ValueError, TypeError):
                velocity = 0
        
        is_long = (length_sensors >= params['LONG_PATTERN_MINIMUM_SENSORS'])
        
        amplitudes = []
        for col_idx in range(13, min(len(row), 50)):
            if (col_idx < len(row) and row[col_idx] and row[col_idx].value is not None and 
                isinstance(row[col_idx].value, (int, float)) and row[col_idx].value > 0):
                amplitudes.append(float(row[col_idx].value))
        
        high_amp_count = count_consecutive_high_amplitude(amplitudes, params['HIGH_AMPLITUDE_MINIMUM_VALUE'])
        is_high_amplitude = (high_amp_count >= params['HIGH_AMPLITUDE_MINIMUM_PATTERN_LENGTH'])
        is_hapc = is_high_amplitude and (direction == 'a') and (length_sensors >= params['HAPC_PATTERN_MINIMUM_SENSORS'])
        is_harpc = is_high_amplitude and (direction == 'r') and (length_sensors >= params['HAPC_PATTERN_MINIMUM_SENSORS'])
        
        pattern_length_category = "Long" if is_long else "Short"
        
        return {
            'length_category': pattern_length_category,
            'direction': direction,
            'velocity': velocity,
            'amplitudes': amplitudes,
            'is_hapc': is_hapc,
            'is_harpc': is_harpc,
            'starting_region': None
        }
        
    except (ValueError, TypeError, IndexError):
        return {
            'length_category': "Short",
            'direction': '',
            'velocity': 0,
            'amplitudes': [],
            'is_hapc': False,
            'is_harpc': False,
            'starting_region': None
        }

def count_consecutive_high_amplitude(amplitudes, threshold):
    """Count consecutive sensors above threshold"""
    if not amplitudes or len(amplitudes) < 3:
        return 0
    
    max_consecutive = 0
    current_consecutive = 0
    
    for amp in amplitudes:
        if isinstance(amp, (int, float)) and amp >= threshold:
            current_consecutive += 1
            max_consecutive = max(max_consecutive, current_consecutive)
        else:
            current_consecutive = 0
    
    return max_consecutive

def determine_starting_region(row, sliders):
    """Determine which colon region a pattern starts in based on first active sensor"""
    try:
        slider_values = getSliderValues(sliders)
        if not slider_values:
            return "Ascending"
        
        regions = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
        
        active_regions = []
        active_sliders = []
        for i, region in enumerate(regions):
            if region not in disabled_sections and i < len(slider_values):
                active_regions.append(region)
                active_sliders.append(slider_values[i])
        
        if not active_regions:
            return "Ascending"
        
        first_active_sensor = None
        for col_idx in range(13, min(len(row), 50)):
            if (col_idx < len(row) and row[col_idx] and row[col_idx].value is not None and 
                isinstance(row[col_idx].value, (int, float)) and row[col_idx].value > 0):
                first_active_sensor = col_idx - 12
                break
        
        if first_active_sensor is None:
            return active_regions[0]
        
        for i, (start_sensor, end_sensor) in enumerate(active_sliders):
            if start_sensor <= first_active_sensor <= end_sensor:
                return active_regions[i]

        if first_active_sensor < active_sliders[0][0]:
            return active_regions[0]

        for i in range(len(active_sliders) - 1, -1, -1):
            if first_active_sensor >= active_sliders[i][0]:
                return active_regions[i]

        return active_regions[0]
        
    except (ValueError, TypeError, IndexError):
        return "Ascending"

def determine_ending_region(row, sliders):
    """Determine which colon region a pattern ends in based on last active sensor"""
    try:
        slider_values = getSliderValues(sliders)
        if not slider_values:
            return "Rectum"
        
        regions = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
        
        active_regions = []
        active_sliders = []
        for i, region in enumerate(regions):
            if region not in disabled_sections and i < len(slider_values):
                active_regions.append(region)
                active_sliders.append(slider_values[i])
        
        if not active_regions:
            return "Rectum"
        
        last_active_sensor = None
        for row_idx in range(len(row) - 1, 12, -1):
            if (row_idx < len(row) and row[row_idx] and row[row_idx].value is not None and 
                isinstance(row[row_idx].value, (int, float)) and row[row_idx].value > 0):
                last_active_sensor = row_idx - 12
                break
        
        if last_active_sensor is None:
            return active_regions[-1]
        
        for i, (start_sensor, end_sensor) in enumerate(active_sliders):
            if start_sensor <= last_active_sensor <= end_sensor:
                return active_regions[i]

        if last_active_sensor > active_sliders[-1][1]:
            return active_regions[-1]

        for i in range(len(active_sliders)):
            if last_active_sensor <= active_sliders[i][1]:
                return active_regions[i]

        return active_regions[-1]
        
    except (ValueError, TypeError, IndexError):
        return "Rectum"

def custom_sort(item):
    order = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    return order.index(item)

def remove_disabled_sections(section):
    global disabled_sections
    if section in disabled_sections:
        disabled_sections.remove(section)
    disabled_sections = sorted(disabled_sections, key=custom_sort)

def add_disabled_sections(section):
    global disabled_sections
    if section not in disabled_sections:
        disabled_sections.append(section)
    disabled_sections = sorted(disabled_sections, key=custom_sort)

def reset_disabled_sections():
    global disabled_sections
    disabled_sections = []

def exportToXlsx(data, file_name, sliders, events, settings_sliders, pattern_params=None, first_event_field=None, original_first_row=None):
    try:
        base_name, ext = file_name.rsplit('.', 1)
        new_file_name = f"{base_name}_analysis.xlsx"

        data.to_excel(new_file_name, index=False)

        # Restore original first row if provided
        if original_first_row:
            wb = load_workbook(new_file_name)
            ws = wb.active
            
            for col in range(1, ws.max_column + 1):
                ws.cell(row=1, column=col).value = None
            
            for i, value in enumerate(original_first_row):
                if value is not None:
                    target_col = i + 1
                    if target_col >= 12:
                        target_col += 2
                    ws.cell(row=1, column=target_col).value = value
            
            wb.save(new_file_name)

        insertEmptyRows(new_file_name, 12)
        create_space_for_comprehensive_table(new_file_name)
        mergeAndColorCells(new_file_name, sliders)

        # Add first event at 0-0-0
        from exportToExcelScreen.events import get_first_event_name
        first_event_name = get_first_event_name()
        addEventNameAtGivenTime(new_file_name, 0, 0, 0, first_event_name)

        event_names = [first_event_name]
        for time, event_name in events.items():
            event_names.append(event_name)
            try:
                total_seconds = time // 10
                hour, remainder = divmod(total_seconds, 3600)
                minute, second = divmod(remainder, 60)
                addEventNameAtGivenTime(new_file_name, hour, minute, second, event_name)
            except Exception as e:
                print(f"Error processing event {event_name}: {e}")
                raise
        
        wb = assignSectionsBasedOnStartSection(new_file_name, sliders, event_names, settings_sliders, pattern_params)
        
        file_name = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], 
            initialfile=new_file_name
        )
        
        wb.save(file_name)
        print(f"Data successfully exported to {new_file_name}")
        
    except Exception as e:
        print(f"Error exporting data to Excel: {e}")

def getSliderValues(sliders):
    list_with_slider_tuples = []
    for i in range(len(sliders)):
        value1 = -1
        value2 = -1
        try:
            slider_values = sliders[i].get()
            if slider_values is None:
                continue
                
            for element in slider_values:
                if value1 == -1:
                    value1 = round(element) if element is not None else 1
                else:
                    value2 = round(element) if element is not None else 1
        except Exception:
            value1, value2 = 1, 10
            
        slider_tuple = (value1, value2)
        list_with_slider_tuples.append(slider_tuple)
    return list_with_slider_tuples

def mergeAndColorCells(file_name, sliders):
    wb = load_workbook(file_name)
    ws = wb.active
    
    sections = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    sliders = getSliderValues(sliders)

    for section in disabled_sections:
        sections.remove(section)
        sliders.pop(0)

    colors = {
        "Ascending": "A9D08E",
        "Transverse": "BDD7EE",
        "Descending": "F8CBAD",
        "Sigmoid": "D9D9D9",
        "Rectum": "B1A0C7"
    }

    # Insert columns for End Region and spacing
    ws.insert_cols(12)
    ws.insert_cols(13)

    ws.cell(row=72, column=12, value="End Region")

    # Color sequence table headers
    header_color = "BFBFBF" 
    for col in range(1, 13):
        header_cell = ws.cell(row=72, column=col)
        if header_cell.value:
            header_cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Region headers and sensor numbers
    for i, (start, end) in enumerate(sliders):
        start_col = get_column_letter(start + 13)
        end_col = get_column_letter(end + 13)
        
        # Region headers on row 71
        ws.merge_cells(f'{start_col}71:{end_col}71')
        cell = ws[f'{start_col}71']
        cell.value = sections[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(start_color=colors[sections[i]], end_color=colors[sections[i]], fill_type="solid")
        cell.fill = fill

        for col in range(start + 13, end + 14):
            ws[f'{get_column_letter(col)}71'].fill = fill

        # Sensor numbers on row 72
        for sensor_num in range(start, end + 1):
            sensor_col = sensor_num + 13
            sensor_cell = ws.cell(row=72, column=sensor_col, value=sensor_num)
            sensor_cell.alignment = Alignment(horizontal='center', vertical='center')
            sensor_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    wb.save(file_name)

def addEventNameAtGivenTime(file_name, hour, minute, second, event_name):
    wb = load_workbook(file_name)
    ws = wb.active
    
    insertion_row = None
    for row in range(27, ws.max_row + 1):
        try:
            cell_hour = ws.cell(row=row, column=2).value
            cell_minute = ws.cell(row=row, column=3).value
            cell_second = ws.cell(row=row, column=4).value
            
            if cell_hour is None or cell_minute is None or cell_second is None:
                continue
            
            try:
                cell_hour_int = int(cell_hour)
                cell_minute_int = int(cell_minute)
                cell_second_int = int(cell_second)
                hour_int = int(hour)
                minute_int = int(minute)
                second_int = int(second)
            except (ValueError, TypeError):
                continue
            
            if (cell_hour_int > hour_int or 
                (cell_hour_int == hour_int and cell_minute_int > minute_int) or 
                (cell_hour_int == hour_int and cell_minute_int == minute_int and cell_second_int >= second_int)):
                insertion_row = row
                break
                
        except Exception:
            continue
    
    if insertion_row is None:
        insertion_row = ws.max_row + 1
    
    ws.insert_rows(insertion_row)
    
    for col in range(1, 13):
        cell = ws.cell(row=insertion_row, column=col)
        cell.value = event_name
        cell.fill = PatternFill(start_color=EVENT_COLOR, end_color=EVENT_COLOR, fill_type="solid")
    
    ws.cell(row=insertion_row, column=2, value=hour)
    ws.cell(row=insertion_row, column=3, value=minute)
    ws.cell(row=insertion_row, column=4, value=second)

    wb.save(file_name)

def insertEmptyRows(file_name, amount):
    wb = load_workbook(file_name)
    ws = wb.active

    for i in range(amount):
        ws.insert_rows(13 + i)
    
    wb.save(file_name)

def assignSectionsBasedOnStartSection(file_name, sliders, event_names, settings_sliders, pattern_params=None):
    params = get_pattern_parameters(pattern_params) if pattern_params else {
        'LONG_PATTERN_MINIMUM_SENSORS': 5,
        'HAPC_PATTERN_MINIMUM_SENSORS': 5,
        'HIGH_AMPLITUDE_MINIMUM_PATTERN_LENGTH': 3,
        'HIGH_AMPLITUDE_MINIMUM_VALUE': 100
    }

    try:
        if settings_sliders and len(settings_sliders) > 0:
            slider_value = settings_sliders[0].get()
            distance_between_sensors = int(round(slider_value)) if slider_value is not None else 25
        else:
            distance_between_sensors = 25
    except Exception:
        distance_between_sensors = 25
    
    wb = load_workbook(file_name)
    ws = wb.active

    all_events = event_names.copy()

    comprehensive_stats = initialize_comprehensive_statistics(all_events)
    
    sections = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    colors = {
        "Ascending": "A9D08E",
        "Transverse": "BDD7EE", 
        "Descending": "F8CBAD",
        "Sigmoid": "D9D9D9",
        "Rectum": "B1A0C7",
        "Ascending tot in Rectum": "81BA5A",
        "Transverse tot in Rectum": "81B2DF",
        "Descending tot in Rectum": "F2A16A",
        "Sigmoid tot in Rectum": "BEBEBE"
    }

    slider_values = getSliderValues(sliders)
    
    counters = {}
    length_counters = {}
    high_amplitude_counters = {}
    
    # Initialize with first event from event_names if any exist
    if event_names:
        current_event = event_names[0]
    else:
        current_event = "Default"  # Fallback if no events defined
    
    for event_name in event_names:
        counters[event_name] = {}
        length_counters[event_name] = {}  
        high_amplitude_counters[event_name] = {}

    counter_template = {
        "Ascending": 0, "Transverse": 0, "Descending": 0, "Sigmoid": 0, "Rectum": 0,
        "Ascending tot in Rectum": 0, "Transverse tot in Rectum": 0,
        "Descending tot in Rectum": 0, "Sigmoid tot in Rectum": 0,
    }

    length_counter_template = {
        "Long s": 0, "Short s": 0, "Long r": 0, "Short r": 0, "Long a": 0, "Short a": 0,
        "cyclic s": 0, "cyclic r": 0, "cyclic a": 0,
    }

    high_amplitude_counters_template = {"HAPCs": 0, "HARPCs": 0}

    # Remove disabled sections
    keys_to_remove = []
    for section in disabled_sections:
        if section in sections:
            sections.remove(section)
            slider_values.pop(0)
        for key in counter_template.keys():
            if section in key:
                keys_to_remove.append(key)

    for key in keys_to_remove:
        if key in counter_template:
            del counter_template[key]

    counter = counter_template.copy()
    length_counter = length_counter_template.copy()
    high_amplitude_counter = high_amplitude_counters_template.copy()
    
    if event_names:
        current_event = event_names[0]
    else:
        current_event = "Default"

    # Process each pattern row
    for row_idx in range(27, ws.max_row + 1):
        try:
            row = [ws.cell(row=row_idx, column=col) for col in range(1, min(ws.max_column + 1, 50))]
            
            if len(row) < 12 or not row[0].value:
                continue

            # Check for event markers
            if (isinstance(row[0].value, str) and 
                not str(row[0].value).isdigit() and 
                row[0].value not in ['Sequence', 'Hour', 'Minute', 'Second', 'Sample']):
                
                for event in all_events:
                    if event not in counters or not counters[event]:
                        counters[event] = counter_template.copy()
                        length_counters[event] = length_counter_template.copy() 
                        high_amplitude_counters[event] = high_amplitude_counters_template.copy()

                # Apply anti-double-counting
                length_counter["Long a"] = max(0, length_counter["Long a"] - high_amplitude_counter["HAPCs"])
                length_counter["Long r"] = max(0, length_counter["Long r"] - high_amplitude_counter["HARPCs"])

                counters[current_event] = dict(counter)
                length_counters[current_event] = dict(length_counter)  
                high_amplitude_counters[current_event] = dict(high_amplitude_counter)
                
                new_event = row[0].value.strip()
                
                if new_event in all_events:
                    current_event = new_event
                    counter = counter_template.copy()
                    length_counter = length_counter_template.copy()
                    high_amplitude_counter = high_amplitude_counters_template.copy()
                
                continue

            # Skip header rows
            if (isinstance(row[0].value, str) and 
                row[0].value in ['Sequence', 'Hour', 'Minute', 'Second', 'Sample']):
                continue

            if (len(row) > 10 and row[10].value and 
                isinstance(row[10].value, str) and 
                not str(row[10].value).replace('.', '').replace('-', '').isdigit()):
                continue
                
            classification = classify_pattern_enhanced(row, sliders, distance_between_sensors, params)
            starting_region = determine_starting_region(row, sliders)
            ending_region = determine_ending_region(row, sliders)
            classification['starting_region'] = starting_region

            # Fill region columns
            region_cell = ws.cell(row=row_idx, column=11, value=starting_region)
            end_region_cell = ws.cell(row=row_idx, column=12, value=ending_region)
            
            region_colors = {
                "Ascending": "A9D08E",
                "Transverse": "BDD7EE", 
                "Descending": "F8CBAD",
                "Sigmoid": "D9D9D9",
                "Rectum": "B1A0C7"
            }

            if starting_region in region_colors:
                region_cell.fill = PatternFill(start_color=region_colors[starting_region], 
                                            end_color=region_colors[starting_region], fill_type="solid")
                
            if ending_region in region_colors:
                end_region_cell.fill = PatternFill(start_color=region_colors[ending_region], 
                                            end_color=region_colors[ending_region], fill_type="solid")
            
            update_comprehensive_stats(comprehensive_stats, classification, current_event, row, sliders)
            
            # Handle HAPCs/HARPCs
            if classification['is_hapc']:
                high_amplitude_counter["HAPCs"] += 1
                comprehensive_stats[current_event]['HAPCs']['count'] += 1
                if classification['velocity'] != 0:
                    comprehensive_stats[current_event]['HAPCs']['velocities'].append(classification['velocity'])
                if classification['amplitudes']:
                    comprehensive_stats[current_event]['HAPCs']['amplitudes'].extend(classification['amplitudes'])
                
                # Color sensor cells green
                for row_idx_sensor in range(13, len(row)):
                    if (row[row_idx_sensor] and row[row_idx_sensor].value is not None and 
                        isinstance(row[row_idx_sensor].value, (int, float)) and 
                        row[row_idx_sensor].value > 0):
                        excel_col = row_idx_sensor + 1
                        sensor_cell = ws.cell(row=row_idx, column=excel_col)
                        sensor_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                        
            elif classification['is_harpc']:
                high_amplitude_counter["HARPCs"] += 1
                comprehensive_stats[current_event]['HARPCs']['count'] += 1
                if classification['velocity'] != 0:
                    comprehensive_stats[current_event]['HARPCs']['velocities'].append(classification['velocity'])
                if classification['amplitudes']:
                    comprehensive_stats[current_event]['HARPCs']['amplitudes'].extend(classification['amplitudes'])
                
                # Color sensor cells red
                for row_idx_sensor in range(13, len(row)):
                    if (row[row_idx_sensor] and row[row_idx_sensor].value is not None and 
                        isinstance(row[row_idx_sensor].value, (int, float)) and 
                        row[row_idx_sensor].value > 0):
                        excel_col = row_idx_sensor + 1
                        sensor_cell = ws.cell(row=row_idx, column=excel_col)
                        sensor_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # Update pattern counters
            pattern = classification['direction']
            length_sensors = 0
            if len(row) > 9 and row[9] and row[9].value is not None:
                try:
                    length_sensors = int(float(str(row[9].value)))
                except (ValueError, TypeError):
                    continue
            
            if pattern and classification['length_category']:
                pattern_type = classification['length_category']
                counter_key = f"{pattern_type} {pattern}"
                
                if counter_key in length_counter:
                    length_counter[counter_key] += 1

            # Regional counting
            if starting_region and starting_region in counter:
                counter[starting_region] += 1
                
                if is_pan_colonic_pattern(row, sliders, starting_region):
                    pan_colonic_key = f'{starting_region} tot in Rectum'
                    if pan_colonic_key in counter:
                        counter[pan_colonic_key] += 1
        except Exception:
            continue

    # Handle final event
    if counters[current_event] == {}:
        for section in sections[:-1] if len(sections) > 1 else []:
            section_key = f'{section} tot in Rectum'
            if section_key in counter:
                counter[section] = max(0, counter[section] - counter[section_key])

        length_counter["Long a"] = max(0, length_counter["Long a"] - high_amplitude_counter["HAPCs"])
        length_counter["Long r"] = max(0, length_counter["Long r"] - high_amplitude_counter["HARPCs"])

        counters[current_event] = counter.copy()
        length_counters[current_event] = length_counter.copy()
        high_amplitude_counters[current_event] = high_amplitude_counter.copy()

    if current_event and (current_event not in counters or all(v == 0 for v in counters[current_event].values())):
        length_counter["Long a"] = max(0, length_counter["Long a"] - high_amplitude_counter["HAPCs"])
        length_counter["Long r"] = max(0, length_counter["Long r"] - high_amplitude_counter["HARPCs"])
        
        counters[current_event] = counter.copy()
        length_counters[current_event] = length_counter.copy()
        high_amplitude_counters[current_event] = high_amplitude_counter.copy()

    # Calculate correct totals before applying HAPC/HARPC corrections
    calculate_correct_totals(comprehensive_stats)
    apply_hapc_harpc_corrections_fixed(comprehensive_stats)

    sync_old_table_with_comprehensive_totals(length_counters, high_amplitude_counters, comprehensive_stats, all_events)
    
    create_comprehensive_analysis_table(wb, comprehensive_stats, all_events)
    
    # Create summary table
    row = 3
    for section in counter_template.keys():
        ws.cell(row=row, column=19, value=section)
        fill = PatternFill(start_color=colors[section], end_color=colors[section], fill_type="solid")
        ws.cell(row=row, column=19).fill = fill
        row += 1
    row += 1
    
    length_counter_row_start = row
    for pattern in list(chain(length_counter_template.keys(), high_amplitude_counters_template.keys())):
        ws.cell(row=row, column=19, value=pattern)
        
        if pattern == "HAPCs":
            fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        elif pattern == "HARPCs":
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        else:
            fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        
        ws.cell(row=row, column=19).fill = fill
        row += 1
    
    # Fill counters
    column = 20
    for (event,value) in counters.items():
        row = 2
        ws.cell(row=row, column=column, value=event)
        fill = PatternFill(start_color=EVENT_COLOR, end_color=EVENT_COLOR, fill_type="solid")
        ws.cell(row=row, column=column).fill = fill
        row += 1
        for (section,value) in value.items():
            ws.cell(row=row, column=column, value=value)
            fill = PatternFill(start_color=colors[section], end_color=colors[section], fill_type="solid")
            ws.cell(row=row, column=column).fill = fill
            row += 1
        column += 1
    
    column = 20
    for (event,value) in length_counters.items():
        row = length_counter_row_start
        for pattern in value.values():
            ws.cell(row=row, column=column, value=pattern)
            row += 1
        column += 1
    
    column = 20
    high_amplitude_start = row
    for (event,value) in high_amplitude_counters.items():
        row = high_amplitude_start
        for pattern in value.values():
            ws.cell(row=row, column=column, value=pattern)
            row += 1
        column += 1
    
    return wb

def calculate_correct_totals(comprehensive_stats):
    """Calculate totals by summing individual region counts"""
    regions = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    region_ranges = ["Ascending - Rectum", "Transverse - Rectum", "Descending - Rectum", "Sigmoid-Rectum"]
    
    for event in comprehensive_stats:
        for pattern_type in comprehensive_stats[event]:
            if pattern_type in ['HAPCs', 'HARPCs', 'cyclic s', 'cyclic r', 'cyclic a']:
                continue  # Skip special categories
                
            # Sum all individual regions and pan-colonic patterns
            total_count = 0
            total_velocities = []
            total_amplitudes = []
            
            for region in regions:
                if region in comprehensive_stats[event][pattern_type]:
                    stats = comprehensive_stats[event][pattern_type][region]
                    total_count += stats['count']
                    total_velocities.extend(stats['velocities'])
                    total_amplitudes.extend(stats['amplitudes'])
            
            for region_range in region_ranges:
                if region_range in comprehensive_stats[event][pattern_type]:
                    stats = comprehensive_stats[event][pattern_type][region_range]
                    total_count += stats['count']
                    total_velocities.extend(stats['velocities'])
                    total_amplitudes.extend(stats['amplitudes'])
            
            # Update the Total
            comprehensive_stats[event][pattern_type]['Total'] = {
                'count': total_count,
                'velocities': total_velocities,
                'amplitudes': total_amplitudes
            }

def apply_hapc_harpc_corrections_fixed(comprehensive_stats):
    """Apply HAPC/HARPC corrections only to the patterns that were actually HAPCs/HARPCs"""
    for event in comprehensive_stats:
        hapc_count = comprehensive_stats[event]['HAPCs']['count']
        harpc_count = comprehensive_stats[event]['HARPCs']['count']
        
        # Only subtract HAPCs from Long a total if there are actual HAPCs
        if hapc_count > 0 and 'Long a' in comprehensive_stats[event]:
            current_total = comprehensive_stats[event]['Long a']['Total']['count']
            comprehensive_stats[event]['Long a']['Total']['count'] = max(0, current_total - hapc_count)
        
        # Only subtract HARPCs from Long r total if there are actual HARPCs  
        if harpc_count > 0 and 'Long r' in comprehensive_stats[event]:
            current_total = comprehensive_stats[event]['Long r']['Total']['count']
            comprehensive_stats[event]['Long r']['Total']['count'] = max(0, current_total - harpc_count)

def sync_old_table_with_comprehensive_totals(length_counters, high_amplitude_counters, comprehensive_stats, all_events):
    """Sync the old table totals with the corrected comprehensive table totals"""
    pattern_mapping = {
        "Long s": "Long s",
        "Short s": "Short s", 
        "Long r": "Long r",
        "Short r": "Short r",
        "Long a": "Long a",
        "Short a": "Short a"
    }
    
    for event in all_events:
        if event in comprehensive_stats and event in length_counters:
            for old_pattern, comp_pattern in pattern_mapping.items():
                if comp_pattern in comprehensive_stats[event] and 'Total' in comprehensive_stats[event][comp_pattern]:
                    # Use the corrected total from comprehensive table
                    corrected_total = comprehensive_stats[event][comp_pattern]['Total']['count']
                    length_counters[event][old_pattern] = corrected_total

def update_comprehensive_stats(comprehensive_stats, classification, current_event, row, sliders):
    """Update comprehensive statistics with pattern data"""
    if current_event not in comprehensive_stats:
        return
    
    pattern_key = f"{classification['length_category']} {classification['direction']}"
    starting_region = classification['starting_region']
    
    if pattern_key not in comprehensive_stats[current_event]:
        return
    
    # Update regional stats
    if starting_region in comprehensive_stats[current_event][pattern_key]:
        stats = comprehensive_stats[current_event][pattern_key][starting_region]
        stats['count'] += 1
        if classification['velocity'] != 0:
            stats['velocities'].append(classification['velocity'])
        if classification['amplitudes']:
            stats['amplitudes'].extend(classification['amplitudes'])
    
    # Check for pan-colonic patterns
    if is_pan_colonic_pattern(row, sliders, starting_region):
        if starting_region == "Sigmoid":
            region_range = "Sigmoid-Rectum"
        else:
            region_range = f"{starting_region} - Rectum"
            
        if region_range in comprehensive_stats[current_event][pattern_key]:
            stats = comprehensive_stats[current_event][pattern_key][region_range]
            stats['count'] += 1
            if classification['velocity'] != 0:
                stats['velocities'].append(classification['velocity'])
            if classification['amplitudes']:
                stats['amplitudes'].extend(classification['amplitudes'])

def is_pan_colonic_pattern(row, sliders, starting_region):
    """Check if pattern propagates to rectum"""
    slider_values = getSliderValues(sliders)
    if len(slider_values) < 5:
        return False
    
    rectum_start, rectum_end = slider_values[4]
    
    # Check rectum activity
    has_rectum_activity = False
    for col in range(13, min(len(row), rectum_end + 13)):
        sensor_num = col - 12
        if (rectum_start <= sensor_num <= rectum_end and 
            col < len(row) and row[col].value is not None and 
            isinstance(row[col].value, (int, float)) and row[col].value > 0):
            has_rectum_activity = True
            break
    
    # Check starting region activity
    starting_region_idx = None
    regions = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    if starting_region in regions:
        starting_region_idx = regions.index(starting_region)
    
    if starting_region_idx is not None and starting_region_idx < len(slider_values):
        start_sensors, end_sensors = slider_values[starting_region_idx]
        has_starting_activity = any(
            (start_sensors <= (col - 12) <= end_sensors and
             col < len(row) and row[col].value is not None and 
             isinstance(row[col].value, (int, float)) and row[col].value > 0)
            for col in range(13, min(len(row), end_sensors + 13))
        )
        
        return has_rectum_activity and has_starting_activity
    
    return False

def create_space_for_comprehensive_table(file_name):
    """Move existing sequence/pan-colonic tables down to make space for comprehensive table"""
    wb = load_workbook(file_name)
    ws = wb.active
    
    rows_to_insert = 71 - 25
    ws.insert_rows(25, rows_to_insert)
    
    wb.save(file_name)

def create_comprehensive_analysis_table(wb, comprehensive_stats, event_names):
    """Create the comprehensive analysis table at AA2:BC68"""
    ws = wb.active
    
    start_row = 2
    start_col = 27

    ranges_to_unmerge = []
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row >= start_row and merged_range.max_row <= 68 and
            merged_range.min_col >= start_col and merged_range.max_col <= start_col + 30):
            ranges_to_unmerge.append(merged_range)

    for range_to_unmerge in ranges_to_unmerge:
        ws.unmerge_cells(str(range_to_unmerge))
    
    create_comprehensive_table_headers(ws, start_row, start_col, event_names)
    
    current_row = start_row + 2
    
    pattern_types = ["Long s", "Short s", "Long r", "Short r", "Long a", "Short a"]
    regions = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    region_ranges = ["Ascending - Rectum", "Transverse - Rectum", 
                    "Descending - Rectum", "Sigmoid-Rectum"]
    
    for pattern_type in pattern_types:
        for region in regions:
            row_label = f"{pattern_type} {region}"
            create_pattern_row(ws, current_row, start_col, row_label, 
                             comprehensive_stats, event_names, pattern_type, region)
            current_row += 1
        
        for region_range in region_ranges:
            row_label = f"{pattern_type} {region_range}"
            create_pattern_row(ws, current_row, start_col, row_label, 
                             comprehensive_stats, event_names, pattern_type, region_range)
            current_row += 1
        
        row_label = f"Total {pattern_type}"
        create_pattern_row(ws, current_row, start_col, row_label, 
                         comprehensive_stats, event_names, pattern_type, 'Total')
        current_row += 1
    
    for cyclic_type in ['cyclic s', 'cyclic r', 'cyclic a']:
        create_pattern_row(ws, current_row, start_col, cyclic_type, 
                         comprehensive_stats, event_names, cyclic_type, None)
        current_row += 1
    
    create_pattern_row(ws, current_row, start_col, "HAPCs", 
                     comprehensive_stats, event_names, 'HAPCs', None)
    current_row += 1
    create_pattern_row(ws, current_row, start_col, "HARPCs", 
                     comprehensive_stats, event_names, 'HARPCs', None)
    
    apply_comprehensive_table_formatting(ws, start_row, start_col, event_names)

def create_comprehensive_table_headers(ws, start_row, start_col, event_names):
    """Create headers for the comprehensive table"""
    col_offset = 1
    for i, event in enumerate(event_names):
        event_start_col = start_col + col_offset
        
        ws.merge_cells(start_row=start_row, start_column=event_start_col, 
                      end_row=start_row, end_column=event_start_col + 6)
        
        ws.cell(row=start_row, column=event_start_col, value=event)
        col_offset += 7
    
    ws.cell(row=start_row + 1, column=start_col, value="Type of event")
    
    col_offset = 1
    metric_headers = ["Number of patterns", "min vel", "max vel", "mean vel", 
                     "min amp", "max amp", "mean amp"]
    
    for event in event_names:
        for j, metric in enumerate(metric_headers):
            ws.cell(row=start_row + 1, column=start_col + col_offset + j, value=metric)
        col_offset += 7

def create_pattern_row(ws, row, start_col, row_label, comprehensive_stats, 
                      event_names, pattern_type, region_key):
    """Create a single pattern row with all metrics"""
    try:
        cell_ref = f"{get_column_letter(start_col)}{row}"
        if any(cell_ref in merged_range for merged_range in ws.merged_cells.ranges):
            return
        
        ws.cell(row=row, column=start_col, value=row_label)
        
        col_offset = 1
        for event in event_names:
            stats = get_pattern_statistics(comprehensive_stats, event, pattern_type, region_key)
            
            ws.cell(row=row, column=start_col + col_offset, value=stats['count'])
            
            if stats['velocities']:
                ws.cell(row=row, column=start_col + col_offset + 1, 
                       value=round(min(stats['velocities']), 2))
                ws.cell(row=row, column=start_col + col_offset + 2, 
                       value=round(max(stats['velocities']), 2))
                ws.cell(row=row, column=start_col + col_offset + 3, 
                       value=round(sum(stats['velocities']) / len(stats['velocities']), 2))
            
            if stats['amplitudes']:
                ws.cell(row=row, column=start_col + col_offset + 4, 
                       value=round(min(stats['amplitudes']), 1))
                ws.cell(row=row, column=start_col + col_offset + 5, 
                       value=round(max(stats['amplitudes']), 1))
                ws.cell(row=row, column=start_col + col_offset + 6, 
                       value=round(sum(stats['amplitudes']) / len(stats['amplitudes']), 1))
            
            col_offset += 7
            
    except Exception:
        pass

def get_pattern_statistics(comprehensive_stats, event, pattern_type, region_key):
    """Get statistics for a specific pattern type, event, and region"""
    default_stats = {'count': 0, 'velocities': [], 'amplitudes': []}
    
    if event not in comprehensive_stats:
        return default_stats
    
    if pattern_type in ['HAPCs', 'HARPCs', 'cyclic s', 'cyclic r', 'cyclic a']:
        if pattern_type in comprehensive_stats[event]:
            return comprehensive_stats[event][pattern_type]
        else:
            return default_stats
    
    if pattern_type not in comprehensive_stats[event]:
        return default_stats
    
    if region_key is None or region_key not in comprehensive_stats[event][pattern_type]:
        return default_stats
    
    return comprehensive_stats[event][pattern_type][region_key]

def apply_comprehensive_table_formatting(ws, start_row, start_col, event_names):
    """Apply colors and formatting to the comprehensive table"""
    colors = {
        "Ascending": "A9D08E",
        "Transverse": "BDD7EE", 
        "Descending": "F8CBAD",
        "Sigmoid": "D9D9D9",
        "Rectum": "B1A0C7",
        "Ascending - Rectum": "81BA5A", 
        "Transverse - Rectum": "81B2DF",
        "Descending - Rectum": "F2A16A",
        "Sigmoid-Rectum": "BEBEBE"
    }
    
    # Event header colors
    col_offset = 1
    for i, event in enumerate(event_names):
        event_color = ALTERNATING_EVENT_COLORS[i % 2]
        event_start_col = start_col + col_offset
        
        for col in range(event_start_col, event_start_col + 7):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = PatternFill(start_color=event_color, end_color=event_color, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        col_offset += 7
    
    # Metric header colors
    metric_gray = "BFBFBF"
    for col in range(start_col, start_col + 1 + len(event_names) * 7):
        cell = ws.cell(row=start_row + 1, column=col)
        cell.fill = PatternFill(start_color=metric_gray, end_color=metric_gray, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Pattern type row colors
    current_row = start_row + 2
    pattern_types = ["Long s", "Short s", "Long r", "Short r", "Long a", "Short a"]
    regions = ["Ascending", "Transverse", "Descending", "Sigmoid", "Rectum"]
    region_ranges = ["Ascending - Rectum", "Transverse - Rectum", 
                    "Descending - Rectum", "Sigmoid-Rectum"]
    
    for pattern_type in pattern_types:
        for region in regions:
            if region in colors:
                cell = ws.cell(row=current_row, column=start_col)
                cell.fill = PatternFill(start_color=colors[region], end_color=colors[region], fill_type="solid")
            current_row += 1
        
        for region_range in region_ranges:
            if region_range in colors:
                cell = ws.cell(row=current_row, column=start_col)
                cell.fill = PatternFill(start_color=colors[region_range], end_color=colors[region_range], fill_type="solid")
            current_row += 1
        
        current_row += 1
    
    current_row += 5

    special_pattern_rows = {
        64: ("cyclic s", "F4B084"),
        65: ("cyclic r", "F4B084"),  
        66: ("cyclic a", "F4B084"),
        67: ("HAPCs", "92D050"),
        68: ("HARPCs", "FF0000")
    }

    for row_num, (pattern_type, color) in special_pattern_rows.items():
        cell = ws.cell(row=row_num, column=start_col)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")